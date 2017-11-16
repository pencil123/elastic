#!/usr/bin/python
#-*- coding: UTF-8 -*-
import datetime
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,colors,fills

class Execl(object):
	def __init__(self):
		self.wb = Workbook()
		self.string_filename = ''

	def create(self,data_list):
		main_list = []
		agent_list = []
		dev_list = []
		other_list = []

		all_ws = self.wb.get_active_sheet()
		all_ws.title = 'all'
		i = 1
		doc_count = len(data_list)
		for num in range(len(data_list)):
			i += 1
			all_ws.cell(row=i,column=1).value=data_list[num]['domain']
			all_ws.cell(row=i,column=2).value=data_list[num]['uv']
			str_percent = "=B" + str(i) + "/$B$" + str(doc_count+2)
			all_ws.cell(row=i,column=3).value= str_percent
			all_ws.cell(row=i,column=3).number_format = '0.0000%'

			if data_list[num]['user'] == 1:
				dev_list.append(data_list[num])
			if data_list[num]['user'] == 2:
				agent_list.append(data_list[num])
			if data_list[num]['user'] == 3:
				main_list.append(data_list[num])
			if data_list[num]['user'] == 4:
				other_list.append(data_list[num])

		self.edit_head(all_ws)
		self.edit_bottom(all_ws,doc_count)

		self.create_main(main_list)
		self.create_agent(agent_list)
		self.create_dev(dev_list)
		self.create_other(other_list)


	def create_main(self,data_list):
		main_ws = self.wb.create_sheet(title='Official')
		i = 1
		doc_count = len(data_list)
		for num in range(len(data_list)):
			i += 1
			main_ws.cell(row=i,column=1).value=data_list[num]['domain']
			main_ws.cell(row=i,column=2).value=data_list[num]['uv']
			str_percent = "=B" + str(i) + "/$B$" + str(doc_count+2)
			main_ws.cell(row=i,column=3).value= str_percent
			main_ws.cell(row=i,column=3).number_format = '0.0000%'

		self.edit_head(main_ws)
		self.edit_bottom(main_ws,doc_count)

	def create_agent(self,data_list):
		agent_ws = self.wb.create_sheet(title='Agent')
		i = 1
		doc_count = len(data_list)
		for num in range(len(data_list)):

			i += 1
			agent_ws.cell(row=i,column=1).value=data_list[num]['domain']
			agent_ws.cell(row=i,column=2).value=data_list[num]['uv']
			str_percent = "=B" + str(i) + "/$B$" + str(doc_count+2)
			agent_ws.cell(row=i,column=3).value= str_percent
			agent_ws.cell(row=i,column=3).number_format = '0.0000%'

		self.edit_head(agent_ws)
		self.edit_bottom(agent_ws,doc_count)

	def create_dev(self,data_list):
		dev_ws = self.wb.create_sheet(title='developer')
		i = 1
		doc_count = len(data_list)
		for num in range(len(data_list)):

			i += 1
			dev_ws.cell(row=i,column=1).value=data_list[num]['domain']
			dev_ws.cell(row=i,column=2).value=data_list[num]['uv']
			str_percent = "=B" + str(i) + "/$B$" + str(doc_count+2)
			dev_ws.cell(row=i,column=3).value= str_percent
			dev_ws.cell(row=i,column=3).number_format = '0.0000%'

		self.edit_head(dev_ws)
		self.edit_bottom(dev_ws,doc_count)

	def create_other(self,data_list):
		other_ws = self.wb.create_sheet(title='other')
		i = 1
		doc_count = len(data_list)
		for num in range(len(data_list)):

			i += 1
			other_ws.cell(row=i,column=1).value=data_list[num]['domain']
			other_ws.cell(row=i,column=2).value=data_list[num]['uv']
			str_percent = "=B" + str(i) + "/$B$" + str(doc_count+2)
			other_ws.cell(row=i,column=3).value= str_percent
			other_ws.cell(row=i,column=3).number_format = '0.0000%'

		self.edit_head(other_ws)
		self.edit_bottom(other_ws,doc_count)
		

	def edit_head(self,ws):
		ws.cell(row=1,column=1).value = "域名".decode('utf-8')
		ws.cell(row=1,column=2).value = "uv"
		ws.cell(row=1,column=3).value = "百分比".decode('utf-8')
		ws.cell(row=1,column=1).font = Font(size=14)
		ws.cell(row=1,column=2).font = Font(size=14)
		ws.cell(row=1,column=3).font = Font(size=14)

		ws.cell(row=1,column=1).fill = PatternFill(start_color=colors.YELLOW, fill_type=fills.FILL_SOLID)
		ws.cell(row=1,column=2).fill = PatternFill(start_color=colors.YELLOW, fill_type=fills.FILL_SOLID)
		ws.cell(row=1,column=3).fill = PatternFill(start_color=colors.YELLOW, fill_type=fills.FILL_SOLID)
		ws.column_dimensions['A'].width = 20
		ws.column_dimensions['B'].width = 20
		ws.column_dimensions['C'].width = 20
	def edit_bottom(self,ws,doc_count):
		ws.cell(row=doc_count+2,column=2).value = "=SUM(B2:B" + str(doc_count+1) + ")"

	def save(self):
		self.string_filename = time.strftime('%Y-%m-%d', time.localtime(time.time())) + ".xlsx"
		self.wb.save(self.string_filename)
		return self.string_filename

	def delete(self):
		os.remove(self.string_filename)